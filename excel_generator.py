"""
excel_generator.py — Fill the Energybae Excel template with extracted bill data.

Takes structured bill data (from extractor.py) and fills it into the provided
Excel template, preserving all solar sizing formulas.
"""

import os
import copy
from datetime import datetime
from openpyxl import load_workbook


# Path to the template file
TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__),
    "Data_provided",
    "Copy of Pranay HOME E-Bill Analysis.xlsx",
)


def parse_month(month_str: str) -> datetime:
    """
    Parse a month string like '2025-02' or '2025-2' into a datetime.
    
    Args:
        month_str: Month in YYYY-MM format.
        
    Returns:
        datetime object set to the 1st of that month.
    """
    try:
        return datetime.strptime(month_str.strip(), "%Y-%m")
    except ValueError:
        # Try alternative formats
        for fmt in ["%Y-%m-%d", "%m-%Y", "%B %Y", "%b %Y"]:
            try:
                return datetime.strptime(month_str.strip(), fmt)
            except ValueError:
                continue
        raise ValueError(f"Cannot parse month: {month_str}")


def generate_excel(
    bill_data: dict,
    output_path: str,
    solar_panel_wattage: int = 600,
    fixed_charges: float = 130.0,
) -> str:
    """
    Fill the Excel template with extracted bill data.
    
    Args:
        bill_data: Dictionary from extractor.py with consumer info and monthly data.
        output_path: Where to save the generated Excel file.
        solar_panel_wattage: Wattage of solar panels (default 600W).
        fixed_charges: Monthly fixed charges (default ₹130).
        
    Returns:
        Path to the generated Excel file.
    """
    # Load template
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # --- Clear old data (preserve structure and formulas) ---
    # Clear consumer 1 header data
    for cell in ["D1", "D2", "D3", "D4", "D5"]:
        ws[cell].value = None
    # Clear consumer 2 header data
    for cell in ["H1", "H2", "H3", "H4", "H5"]:
        ws[cell].value = None

    # Clear monthly data rows (rows 9-21) for both consumers
    for row in range(9, 22):
        for col in ["B", "C", "D", "E", "G", "H", "I"]:
            ws[f"{col}{row}"].value = None

    # --- Fill in new data (consumer 1 — columns B-F) ---

    # Header fields
    ws["D1"] = bill_data.get("consumer_name", "")
    ws["D2"] = bill_data.get("consumer_no", "")
    ws["D3"] = fixed_charges
    ws["D4"] = bill_data.get("sanctioned_load", "")
    ws["D5"] = bill_data.get("connection_type", "")

    # Solar panel wattage
    ws["C7"] = solar_panel_wattage

    # --- Fill monthly consumption data ---
    monthly_data = bill_data.get("monthly_data", [])

    # Sort by month chronologically
    monthly_data_sorted = sorted(monthly_data, key=lambda x: x.get("month", ""))

    # Fill rows 9 onwards (up to 13 months)
    for i, entry in enumerate(monthly_data_sorted[:13]):
        row = 9 + i
        ws[f"B{row}"] = i + 2  # Sr.No (starting from 2 to match template)

        # Parse and set month
        try:
            month_dt = parse_month(entry["month"])
            ws[f"C{row}"] = month_dt
            ws[f"C{row}"].number_format = "MMM-YYYY"
        except (ValueError, KeyError):
            ws[f"C{row}"] = entry.get("month", "")

        # Set units
        units = entry.get("units", 0)
        ws[f"D{row}"] = units if units is not None else 0

        # Set bill amount if available
        bill_amount = entry.get("bill_amount")
        if bill_amount is not None:
            ws[f"E{row}"] = bill_amount
            # Unit cost formula
            ws[f"F{row}"] = f"=(E{row}-$D$3)/D{row}"

    # If we have the current bill amount and it's the last entry, ensure it's set
    current_amount = bill_data.get("current_bill_amount")
    current_units = bill_data.get("current_units")
    current_month = bill_data.get("current_month")

    if current_month and current_units is not None:
        # Check if current month is already in monthly_data
        current_found = False
        for i, entry in enumerate(monthly_data_sorted[:13]):
            if entry.get("month", "").startswith(current_month[:7] if current_month else ""):
                current_found = True
                row = 9 + i
                ws[f"D{row}"] = current_units
                if current_amount is not None:
                    ws[f"E{row}"] = current_amount
                    ws[f"F{row}"] = f"=(E{row}-$D$3)/D{row}"
                break

        if not current_found:
            # Add as the next available row
            next_row = 9 + len(monthly_data_sorted[:13])
            if next_row <= 21:
                ws[f"B{next_row}"] = len(monthly_data_sorted[:13]) + 2
                try:
                    month_dt = parse_month(current_month)
                    ws[f"C{next_row}"] = month_dt
                    ws[f"C{next_row}"].number_format = "MMM-YYYY"
                except ValueError:
                    ws[f"C{next_row}"] = current_month
                ws[f"D{next_row}"] = current_units
                if current_amount is not None:
                    ws[f"E{next_row}"] = current_amount
                    ws[f"F{next_row}"] = f"=(E{next_row}-$D$3)/D{next_row}"

    # --- Ensure calculation formulas are intact (rows 22-30) ---
    # These should already be preserved from template, but let's be safe
    ws["D22"] = "=AVERAGE(D9:D21)"
    ws["E22"] = "=AVERAGE(E9:E21)"
    ws["F22"] = "=AVERAGE(F9:F21)"
    ws["C22"] = "Average"
    ws["C23"] = "kW"
    ws["D23"] = "=(D22*12*1.1)/1400"
    ws["C24"] = "Solar Panels"
    ws["D24"] = "=D23/$C$7*1000"
    ws["C25"] = "Solar capacity"
    ws["D25"] = "=ROUND(D24,0)*$C$7/1000"
    ws["C26"] = "Number of Panels"
    ws["D26"] = "=D25/$C$7*1000"
    ws["C29"] = "Total solar capacity"
    ws["D29"] = "=SUM(D25,H25)"
    ws["C30"] = "Number of solar panels"
    ws["D30"] = "=SUM(D26,H26)"

    # --- Save output ---
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    return output_path
